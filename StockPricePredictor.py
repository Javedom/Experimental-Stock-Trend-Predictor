import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pickle
import sys
import argparse
import yfinance as yf
from datetime import datetime, timedelta

# -------------------------------------------------------------------------
# CONFIGURATION - Edit these values for your analysis
# -------------------------------------------------------------------------
# List of ticker symbols to analyze
DEFAULT_TICKERS = ['AAPL', 'MSFT', 'GOOG', 'AMZN', 'META', 'TSLA', 'NVDA']

# Date range for historical data
DEFAULT_START_DATE = '2016-01-01'  # Format: YYYY-MM-DD
DEFAULT_END_DATE = datetime.today().strftime('%Y-%m-%d')  # Today's date by default

# Model file path
DEFAULT_MODEL_FILE = 'stock_price_model.pkl'

# -------------------------------------------------------------------------
# Scikit-Learn Imports
# -------------------------------------------------------------------------
from sklearn.model_selection import train_test_split, cross_val_score, GridSearchCV, StratifiedKFold
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics import (
    classification_report,
    precision_score,
    recall_score,
    f1_score,
    confusion_matrix,
    ConfusionMatrixDisplay,
    roc_curve,
    auc,
    accuracy_score
)

# -------------------------------------------------------------------------
# Hyperparameter optimization
# -------------------------------------------------------------------------
import optuna

# -----------------------------------------------------------------------------
# 1) Data Collection Functions
# -----------------------------------------------------------------------------
def fetch_stock_data(tickers, start_date, end_date):
    """
    Fetches stock price data for a list of ticker symbols within a date range.
    
    Arguments:
        tickers: List of ticker symbols
        start_date: Start date in 'YYYY-MM-DD' format
        end_date: End date in 'YYYY-MM-DD' format
        
    Returns:
        Dictionary containing the stock price data for each ticker
    """
    print(f"Fetching stock price data for {len(tickers)} tickers...")
    
    all_data = {}
    
    # Download all tickers at once (more efficient)
    try:
        # Use group="ticker" to separate data by ticker
        stock_data = yf.download(tickers, start=start_date, end=end_date, 
                                interval="1d", auto_adjust=False, group_by="ticker")
        
        # Process each ticker's data
        for ticker in tickers:
            if ticker in stock_data.columns.levels[0]:
                ticker_data = stock_data[ticker].copy()
                all_data[ticker] = ticker_data
                print(f"✓ Fetched data for {ticker}: {len(ticker_data)} data points")
            else:
                print(f"Warning: No data found for {ticker}")
        
    except Exception as e:
        print(f"Error fetching data: {str(e)}")
        # Fall back to fetching tickers individually if the batch request fails
        for ticker in tickers:
            try:
                ticker_data = yf.download(ticker, start=start_date, end=end_date, 
                                        interval="1d", auto_adjust=False)
                
                if ticker_data.empty:
                    print(f"Warning: No data found for {ticker}")
                    continue
                    
                all_data[ticker] = ticker_data
                print(f"✓ Fetched data for {ticker}: {len(ticker_data)} data points")
                
            except Exception as e:
                print(f"Error fetching data for {ticker}: {str(e)}")
    
    print(f"Successfully fetched data for {len(all_data)} out of {len(tickers)} tickers")
    return all_data

def get_quarterly_prices(stock_data):
    """
    Extracts quarterly prices from daily stock data.
    
    Arguments:
        stock_data: Dictionary of DataFrames with daily stock price data
        
    Returns:
        Dictionary of DataFrames with quarterly stock price data
    """
    quarterly_data = {}
    
    for ticker, data in stock_data.items():
        try:
            # Check column structure (handle both single-level and multi-level columns)
            if isinstance(data.columns, pd.MultiIndex):
                # For multi-level index, the data comes from batch download
                print(f"Processing multi-level columns for {ticker}")
                
                # Select the appropriate price column for this ticker
                if ('Adj Close', ticker) in data.columns:
                    price_series = data[('Adj Close', ticker)]
                elif 'Adj Close' in data.columns:
                    price_series = data['Adj Close']
                elif ('Close', ticker) in data.columns:
                    price_series = data[('Close', ticker)]
                else:
                    price_series = data['Close']
            else:
                # For single-level index (from individual downloads)
                if 'Adj Close' in data.columns:
                    price_series = data['Adj Close']
                else:
                    price_series = data['Close']
                    print(f"Note: Using 'Close' instead of 'Adj Close' for {ticker}")
            
            # Resample to end of quarter prices
            quarterly = price_series.resample('QE').last()
            
            # Calculate quarterly returns
            quarterly_returns = quarterly.pct_change()
            
            # Create DataFrame with quarterly prices and returns
            quarterly_df = pd.DataFrame({
                'price': quarterly,
                'quarterly_return': quarterly_returns
            })
            
            # Add trend column
            quarterly_df['trend'] = quarterly_df['quarterly_return'].apply(
                lambda x: 'up' if pd.notna(x) and x > 0 else 'down'
            )
            
            # Calculate next quarter's trend (this will be our target variable)
            quarterly_df['next_quarter_trend'] = quarterly_df['trend'].shift(-1)
            
            quarterly_data[ticker] = quarterly_df
            print(f"✓ Generated quarterly data for {ticker}: {len(quarterly_df)} quarters")
            
        except Exception as e:
            print(f"Error processing quarterly data for {ticker}: {str(e)}")
            if isinstance(data.columns, pd.MultiIndex):
                print(f"MultiIndex columns for {ticker}: {data.columns.tolist()[:3]}...")
            else:
                print(f"Columns for {ticker}: {data.columns.tolist()}")
            # Continue with next ticker instead of crashing
            continue
    
    return quarterly_data

def fetch_financial_ratios(tickers):
    """
    Fetches financial ratios for a list of ticker symbols.
    
    Arguments:
        tickers: List of ticker symbols
        
    Returns:
        Dictionary containing financial info for each ticker
    """
    print(f"Fetching financial ratios for {len(tickers)} tickers...")
    
    financial_data = {}
    for ticker in tickers:
        try:
            # Create ticker object
            stock = yf.Ticker(ticker)
            
            # Get financial information (current snapshot)
            info = stock.info
            
            # Get quarterly financial statements (historical)
            quarterly_financials = stock.quarterly_financials
            quarterly_balance_sheet = stock.quarterly_balance_sheet
            quarterly_cash_flow = stock.quarterly_cashflow
            
            # Store all data
            financial_data[ticker] = {
                'info': info,
                'quarterly_financials': quarterly_financials,
                'quarterly_balance_sheet': quarterly_balance_sheet,
                'quarterly_cash_flow': quarterly_cash_flow
            }
            
            # Print the date range of historical data availability
            if not quarterly_financials.empty:
                fin_dates = [pd.to_datetime(str(date)).strftime('%Y-%m-%d') for date in quarterly_financials.columns]
                print(f"✓ Fetched financial data for {ticker}: Historical quarters from {fin_dates[-1]} to {fin_dates[0]}")
            else:
                print(f"✓ Fetched financial data for {ticker}: No historical statements available")
            
        except Exception as e:
            print(f"Error fetching financial data for {ticker}: {str(e)}")
    
    return financial_data

def find_closest_date(target_date, available_dates):
    """
    Find the closest date to the target date from the available dates.
    Used to match quarterly price data with the closest financial statement date.
    
    Arguments:
        target_date: Target date to find closest match for
        available_dates: List of available dates to search from
        
    Returns:
        Closest date from available_dates, or None if no dates available
    """
    if len(available_dates) == 0:
        return None
    
    # Convert to datetime if needed
    if not isinstance(target_date, pd.Timestamp):
        target_date = pd.Timestamp(target_date)
    
    # Calculate absolute difference between target and all available dates
    date_diffs = {}
    for date in available_dates:
        if isinstance(date, str):
            date = pd.Timestamp(date)
        date_diffs[date] = abs((target_date - date).days)
    
    # Find date with minimum difference
    closest_date = min(date_diffs.items(), key=lambda x: x[1])[0]
    
    # Only return dates that are not too far (within 120 days)
    if date_diffs[closest_date] <= 120:
        return closest_date
    else:
        return None

def extract_quarterly_ratios(financial_data, quarterly_prices):
    """
    Extracts quarterly financial ratios and combines with quarterly price data.
    Uses actual historical quarterly data where available.
    
    Arguments:
        financial_data: Dictionary containing financial info for each ticker
        quarterly_prices: Dictionary of DataFrames with quarterly stock price data
        
    Returns:
        DataFrame with combined quarterly ratios and price trends
    """
    all_quarterly_data = []
    
    for ticker, data in financial_data.items():
        if ticker not in quarterly_prices:
            continue
            
        price_data = quarterly_prices[ticker]
        
        # Get the quarterly financial statements
        quarterly_financials = data.get('quarterly_financials', pd.DataFrame())
        quarterly_balance_sheet = data.get('quarterly_balance_sheet', pd.DataFrame())
        quarterly_cash_flow = data.get('quarterly_cashflow', pd.DataFrame())
        
        # Check if we have actual historical data
        has_quarterly_data = not (quarterly_financials.empty and quarterly_balance_sheet.empty and quarterly_cash_flow.empty)
        
        if has_quarterly_data:
            print(f"Using historical quarterly financial data for {ticker}")
        else:
            print(f"Warning: No historical quarterly financial data for {ticker}, using latest snapshot only")
        
        # Get current info (this is a snapshot, not historical)
        info = data.get('info', {})
        
        # Get quarterly dates from price data
        quarters = price_data.index
        
        for quarter_date in quarters:
            if pd.isna(price_data.loc[quarter_date, 'next_quarter_trend']):
                continue  # Skip if we don't have the next quarter's trend
            
            # Create base quarter data with price and trend information
            quarter_dict = {
                'ticker': ticker,
                'quarter_date': quarter_date,
                'price': price_data.loc[quarter_date, 'price'],
                'quarterly_return': price_data.loc[quarter_date, 'quarterly_return'],
                'trend': price_data.loc[quarter_date, 'trend'],
                'next_quarter_trend': price_data.loc[quarter_date, 'next_quarter_trend'],
                'data_source': 'historical' if has_quarterly_data else 'current_snapshot'
            }
            
            # ----------------------------------------------------------------
            # 1. Extract truly historical financial data from quarterly statements
            # ----------------------------------------------------------------
            if has_quarterly_data:
                # Find the closest financial statement date to this quarter
                closest_financial_date = find_closest_date(quarter_date, quarterly_financials.columns)
                closest_balance_date = find_closest_date(quarter_date, quarterly_balance_sheet.columns)
                closest_cashflow_date = find_closest_date(quarter_date, quarterly_cash_flow.columns)
                
                # Add historical financial data (key indicators that are available)
                # Income Statement Data
                if closest_financial_date is not None:
                    try:
                        # Total Revenue
                        if 'Total Revenue' in quarterly_financials.index:
                            quarter_dict['totalRevenue'] = quarterly_financials.loc['Total Revenue', closest_financial_date]
                        
                        # Gross Profit
                        if 'Gross Profit' in quarterly_financials.index:
                            quarter_dict['grossProfits'] = quarterly_financials.loc['Gross Profit', closest_financial_date]
                        
                        # EBITDA
                        if 'EBITDA' in quarterly_financials.index:
                            quarter_dict['ebitda'] = quarterly_financials.loc['EBITDA', closest_financial_date]
                        
                        # Net Income
                        if 'Net Income' in quarterly_financials.index:
                            quarter_dict['netIncomeToCommon'] = quarterly_financials.loc['Net Income', closest_financial_date]
                        
                        # Calculate margins if possible
                        if 'Total Revenue' in quarterly_financials.index and 'Gross Profit' in quarterly_financials.index:
                            total_revenue = quarterly_financials.loc['Total Revenue', closest_financial_date]
                            gross_profit = quarterly_financials.loc['Gross Profit', closest_financial_date]
                            if total_revenue > 0:
                                quarter_dict['grossMargins'] = gross_profit / total_revenue
                            
                        if 'Total Revenue' in quarterly_financials.index and 'Operating Income' in quarterly_financials.index:
                            total_revenue = quarterly_financials.loc['Total Revenue', closest_financial_date]
                            operating_income = quarterly_financials.loc['Operating Income', closest_financial_date]
                            if total_revenue > 0:
                                quarter_dict['operatingMargins'] = operating_income / total_revenue
                                
                        if 'Total Revenue' in quarterly_financials.index and 'Net Income' in quarterly_financials.index:
                            total_revenue = quarterly_financials.loc['Total Revenue', closest_financial_date]
                            net_income = quarterly_financials.loc['Net Income', closest_financial_date]
                            if total_revenue > 0:
                                quarter_dict['profitMargins'] = net_income / total_revenue
                                
                    except Exception as e:
                        print(f"Error processing income statement data for {ticker}: {e}")
                
                # Balance Sheet Data
                if closest_balance_date is not None:
                    try:
                        # Total Assets
                        if 'Total Assets' in quarterly_balance_sheet.index:
                            quarter_dict['totalAssets'] = quarterly_balance_sheet.loc['Total Assets', closest_balance_date]
                        
                        # Total Liabilities
                        if 'Total Liabilities Net Minority Interest' in quarterly_balance_sheet.index:
                            quarter_dict['totalLiabilities'] = quarterly_balance_sheet.loc['Total Liabilities Net Minority Interest', closest_balance_date]
                        
                        # Total Cash
                        if 'Cash And Cash Equivalents' in quarterly_balance_sheet.index:
                            quarter_dict['totalCash'] = quarterly_balance_sheet.loc['Cash And Cash Equivalents', closest_balance_date]
                        
                        # Total Debt
                        if 'Total Debt' in quarterly_balance_sheet.index:
                            quarter_dict['totalDebt'] = quarterly_balance_sheet.loc['Total Debt', closest_balance_date]
                        
                        # Calculate ratios
                        if 'Total Assets' in quarterly_balance_sheet.index and 'Total Debt' in quarterly_balance_sheet.index:
                            total_assets = quarterly_balance_sheet.loc['Total Assets', closest_balance_date]
                            total_debt = quarterly_balance_sheet.loc['Total Debt', closest_balance_date]
                            if total_assets > 0:
                                quarter_dict['debtRatio'] = total_debt / total_assets
                                quarter_dict['equityRatio'] = 1 - (total_debt / total_assets)
                        
                        if 'Total Liabilities Net Minority Interest' in quarterly_balance_sheet.index and 'Stockholders Equity' in quarterly_balance_sheet.index:
                            total_liabilities = quarterly_balance_sheet.loc['Total Liabilities Net Minority Interest', closest_balance_date]
                            stockholders_equity = quarterly_balance_sheet.loc['Stockholders Equity', closest_balance_date]
                            if stockholders_equity > 0:
                                quarter_dict['debtToEquity'] = total_liabilities / stockholders_equity
                    
                    except Exception as e:
                        print(f"Error processing balance sheet data for {ticker}: {e}")
                
                # Cash Flow Data
                if closest_cashflow_date is not None:
                    try:
                        # Operating Cash Flow
                        if 'Operating Cash Flow' in quarterly_cash_flow.index:
                            quarter_dict['operatingCashflow'] = quarterly_cash_flow.loc['Operating Cash Flow', closest_cashflow_date]
                        
                        # Free Cash Flow
                        if 'Free Cash Flow' in quarterly_cash_flow.index:
                            quarter_dict['freeCashflow'] = quarterly_cash_flow.loc['Free Cash Flow', closest_cashflow_date]
                    
                    except Exception as e:
                        print(f"Error processing cash flow data for {ticker}: {e}")
            
            # ----------------------------------------------------------------
            # 2. For metrics not in quarterly statements, use current snapshot with warning
            # ----------------------------------------------------------------
            # Define all metrics we want from current snapshot 
            # (only use these as a fallback when historical data is unavailable)
            current_snapshot_metrics = [
                # Company Information
                'fullTimeEmployees',
                
                # Financial Ratios (if not calculated from quarterly data)
                'returnOnAssets', 'returnOnEquity', 'quickRatio', 'currentRatio',
                
                # Stock Metrics
                'beta', 'forwardPE', 'trailingPE', 'priceToBook',
                
                # Dividend Information
                'dividendRate', 'dividendYield', 'payoutRatio',
                
                # Valuation Metrics
                'enterpriseValue', 'enterpriseToEbitda', 'pegRatio', 
                
                # Earnings Information
                'trailingEps', 'forwardEps', 'earningsGrowth', 'revenueGrowth',
                
                # Short Interest
                'shortPercentOfFloat', 'shortRatio', 'sharesOutstanding', 'floatShares'
            ]
            
            # Add metrics from current snapshot
            for metric in current_snapshot_metrics:
                if metric in info and metric not in quarter_dict:
                    # Mark these features as coming from current snapshot
                    metric_name = f"{metric}_current"
                    
                    # Handle date fields
                    if metric in ['exDividendDate', 'dividendDate']:
                        if pd.notna(info[metric]) and info[metric] is not None:
                            try:
                                date_val = pd.to_datetime(info[metric])
                                quarter_dict[metric_name] = date_val.timestamp() / (60*60*24)  # Convert to days
                            except:
                                quarter_dict[metric_name] = np.nan
                        else:
                            quarter_dict[metric_name] = np.nan
                    else:
                        quarter_dict[metric_name] = info[metric]
                        
            # ----------------------------------------------------------------
            # 3. Add technical indicators that are truly point-in-time
            # ----------------------------------------------------------------
            # These metrics can be calculated from price data and represent the state at that time
            try:
                # Calculate technical indicators based on price data up to this point
                price_history = price_data.loc[:quarter_date]
                
                # Simple Moving Averages
                if len(price_history) >= 4:  # Need at least 4 quarters for 1-year SMA
                    quarter_dict['sma_4q'] = price_history['price'].rolling(window=4).mean().iloc[-1]
                    quarter_dict['price_to_sma_4q'] = price_history['price'].iloc[-1] / quarter_dict['sma_4q'] if quarter_dict['sma_4q'] != 0 else np.nan
                
                # Momentum (6-month and 12-month)
                if len(price_history) >= 2:
                    quarter_dict['momentum_2q'] = price_history['price'].iloc[-1] / price_history['price'].iloc[-2] - 1 if price_history['price'].iloc[-2] != 0 else np.nan
                
                if len(price_history) >= 4:
                    quarter_dict['momentum_4q'] = price_history['price'].iloc[-1] / price_history['price'].iloc[-4] - 1 if price_history['price'].iloc[-4] != 0 else np.nan
                
                # Volatility (standard deviation of returns over past 4 quarters)
                if len(price_history) >= 4:
                    quarter_dict['volatility_4q'] = price_history['quarterly_return'].tail(4).std()
                
            except Exception as e:
                print(f"Error calculating technical indicators for {ticker}: {e}")
            
            # Add the quarter data to our collection
            all_quarterly_data.append(quarter_dict)
    
    # Convert to DataFrame
    if all_quarterly_data:
        result_df = pd.DataFrame(all_quarterly_data)
        # Print the number of columns (features) extracted
        print(f"Extracted {len(result_df.columns) - 7} financial features")  # -7 to account for non-feature columns
        
        # Calculate the percentage of data from historical vs current
        hist_pct = (result_df['data_source'] == 'historical').mean() * 100 if 'data_source' in result_df else 0
        print(f"Data sources: {hist_pct:.1f}% historical quarterly data, {100-hist_pct:.1f}% current snapshot")
        
        return result_df
    else:
        print("Warning: No quarterly data could be extracted")
        return pd.DataFrame()

# -----------------------------------------------------------------------------
# 2) Model Loading Function
# -----------------------------------------------------------------------------
def load_saved_model(model_file='stock_price_model.pkl'):
    """
    Loads a previously trained model and its components from a pickle file.
    
    Arguments:
        model_file: Path to the saved model file
        
    Returns:
        model: The trained model
        scaler: StandardScaler fitted on training data
        feature_names: List of feature column names
        label_encoder: LabelEncoder for labels (if available)
        success: Boolean indicating if loading was successful
    """
    try:
        with open(model_file, 'rb') as f:
            model_data = pickle.load(f)
            
        # Extract components
        model = model_data['model']
        scaler = model_data.get('scaler')
        feature_names = model_data.get('feature_names', [])
        label_encoder = model_data.get('label_encoder')
        
        print(f"✅ Successfully loaded model from '{model_file}'")
        print(f"Model type: {type(model).__name__}")
        print(f"Number of features: {len(feature_names)}")
        
        if label_encoder is None:
            print("Note: Label encoder not found in saved model, will create a default one")
            # Create a default label encoder
            label_encoder = LabelEncoder()
            label_encoder.fit(['down', 'up'])  # Default classes
        
        return model, scaler, feature_names, label_encoder, True
    except Exception as e:
        print(f"❌ Error loading model: {str(e)}")
        return None, None, None, None, False

# -----------------------------------------------------------------------------
# 3) Data Preprocessing
# -----------------------------------------------------------------------------
def preprocess_data(data_df, predict_df=None):
    """
    Prepares the stock price dataset for model training and prediction.
    
    For training data:
    1. Handle missing values
    2. Encode target variable (up/down) to numeric (1/0)
    3. Split features and target
    
    For prediction data (if provided):
    1. Apply same preprocessing steps as training data
    2. Apply same scaling as training data
    
    Arguments:
        data_df: DataFrame with training data (includes next_quarter_trend)
        predict_df: Optional DataFrame with prediction data (without next_quarter_trend)
    
    Returns:
    - X_train: Feature DataFrame for training
    - y_train: Target Series for training (encoded)
    - X_predict: Feature DataFrame for prediction (or empty DataFrame if not provided/error)
    - feature_cols: List of feature column names
    - label_encoder: Fitted LabelEncoder for trend
    - scaler: Fitted StandardScaler for features
    """
    # Make a copy to avoid modifying the original
    train_data = data_df.copy()
    
    # Drop rows with missing target values
    train_data = train_data.dropna(subset=['next_quarter_trend'])
    
    # Identify non-feature columns
    non_feature_cols = ['ticker', 'quarter_date', 'trend', 'next_quarter_trend', 'next_quarter_trend_encoded', 'data_source']
    
    # Identify all potential feature columns (all numeric columns except non-feature columns)
    all_columns = train_data.columns.tolist()
    feature_cols = [col for col in all_columns if col not in non_feature_cols]
    
    # Filter to keep only numeric columns for features
    numeric_cols = train_data.select_dtypes(include=['number']).columns
    feature_cols = [col for col in feature_cols if col in numeric_cols]
    
    # Print the features being used
    print(f"Using {len(feature_cols)} features for model training")
    
    # Fill missing values with median for each feature
    for feature in feature_cols:
        if feature in train_data.columns:
            median_value = train_data[feature].median()
            train_data[feature] = train_data[feature].fillna(median_value)
    
    # Encode the target variable
    label_encoder = LabelEncoder()
    train_data['next_quarter_trend_encoded'] = label_encoder.fit_transform(train_data['next_quarter_trend'])
    
    # Print mapping for clarity
    trend_mapping = dict(zip(label_encoder.classes_, label_encoder.transform(label_encoder.classes_)))
    print(f"Trend encoding: {trend_mapping}")
    
    # Normalize features
    scaler = StandardScaler()
    train_data[feature_cols] = scaler.fit_transform(train_data[feature_cols])
    
    # Split into X_train and y_train
    X_train = train_data[feature_cols]
    y_train = train_data['next_quarter_trend_encoded']
    
    # Initialize X_predict as an empty DataFrame with the right columns
    # This ensures we never return None for X_predict
    X_predict = pd.DataFrame(columns=feature_cols)
    
    # Process prediction data if provided
    if predict_df is not None:
        try:
            predict_data = predict_df.copy()
            
            # Print debug info
            print(f"\nPreprocessing prediction data with shape: {predict_data.shape}")
            print(f"Prediction data columns: {len(predict_data.columns)}")
            
            # Handle missing feature values
            for feature in feature_cols:
                if feature in predict_data.columns:
                    # Use training data median to fill missing values
                    median_value = train_data[feature].median() if not pd.isna(train_data[feature]).all() else 0
                    predict_data[feature] = predict_data[feature].fillna(median_value)
                else:
                    print(f"Warning: Feature '{feature}' missing in prediction data, setting to 0")
                    predict_data[feature] = 0.0
            
            # Apply the same scaling as training data
            predict_data[feature_cols] = scaler.transform(predict_data[feature_cols])
            
            # Final prediction features
            X_predict = predict_data[feature_cols].copy()
            print(f"Prediction features shape: {X_predict.shape}")
            
        except Exception as e:
            print(f"Error preprocessing prediction data: {str(e)}")
            import traceback
            traceback.print_exc()
            print("Using empty DataFrame for X_predict")
            # X_predict is already initialized as an empty DataFrame
    
    return X_train, y_train, X_predict, feature_cols, label_encoder, scaler

# -----------------------------------------------------------------------------
# 4) Cross-Validation
# -----------------------------------------------------------------------------
def cross_validate_model(X, y, n_splits=5):
    """
    Performs Stratified K-Fold cross-validation on the dataset using a RandomForestClassifier.
    Returns the array of accuracy scores and prints the mean.
    """
    model = RandomForestClassifier(class_weight='balanced', random_state=42)
    
    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    cv_scores = cross_val_score(model, X, y, cv=skf, scoring='accuracy')
    
    print(f"K-Fold CV accuracy scores: {cv_scores}")
    print(f"Mean CV accuracy: {cv_scores.mean():.4f}")
    return cv_scores

# -----------------------------------------------------------------------------
# 5) Hyperparameter Tuning (Grid Search)
# -----------------------------------------------------------------------------
def grid_search_tuning(X, y):
    """
    Performs Grid Search for hyperparameter tuning on a RandomForestClassifier.
    Returns the best model found.
    """
    model = RandomForestClassifier(random_state=42)
    
    # Define parameter grid
    param_grid = {
        'n_estimators': [50, 100, 200],
        'max_depth': [None, 10, 20, 30],
        'min_samples_split': [2, 5, 10],
        'class_weight': ['balanced', None]
    }
    
    grid_search = GridSearchCV(
        estimator=model,
        param_grid=param_grid,
        scoring='f1',  # F1 score is good for binary classification with potential imbalance
        cv=5,
        n_jobs=-1,
        verbose=1
    )
    
    print("Starting Grid Search...")
    grid_search.fit(X, y)
    
    print(f"\nBest Parameters: {grid_search.best_params_}")
    print(f"Best F1 Score: {grid_search.best_score_:.4f}")
    
    best_model = grid_search.best_estimator_
    return best_model

# -----------------------------------------------------------------------------
# 6) Hyperparameter Tuning with Optuna
# -----------------------------------------------------------------------------
def tune_rf_with_optuna(X, y, n_trials=30):
    """
    Uses Optuna to perform Bayesian Optimization on a RandomForestClassifier.
    Tries 'n_trials' different hyperparameter combinations, guided by past results.
    Returns the best model found.
    """
    def objective(trial):
        # Suggest hyperparameters
        n_estimators = trial.suggest_int("n_estimators", 50, 300, step=50)
        max_depth = trial.suggest_categorical("max_depth", [None, 10, 20, 30, 40])
        min_samples_split = trial.suggest_int("min_samples_split", 2, 10)
        min_samples_leaf = trial.suggest_int("min_samples_leaf", 1, 5)
        max_features = trial.suggest_categorical("max_features", ['sqrt', 'log2', None])
        class_weight = trial.suggest_categorical("class_weight", ['balanced', None])
        
        # Create the model
        model = RandomForestClassifier(
            n_estimators=n_estimators,
            max_depth=max_depth,
            min_samples_split=min_samples_split,
            min_samples_leaf=min_samples_leaf,
            max_features=max_features,
            class_weight=class_weight,
            random_state=42
        )
        
        # Cross-validation for scoring
        skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        scores = cross_val_score(model, X, y, cv=skf, scoring='f1', n_jobs=-1)
        return scores.mean()  # maximize F1 score
    
    study = optuna.create_study(direction='maximize')
    study.optimize(objective, n_trials=n_trials)
    
    print(f"\n[Optuna] Number of finished trials: {len(study.trials)}")
    print(f"[Optuna] Best trial parameters: {study.best_trial.params}")
    print(f"[Optuna] Best trial CV F1 score: {study.best_trial.value:.4f}")
    
    # Build final model using the best parameters
    best_params = study.best_params
    best_model = RandomForestClassifier(
        **best_params,
        random_state=42
    )
    best_model.fit(X, y)
    
    return best_model

# -----------------------------------------------------------------------------
# 7) Train Model
# -----------------------------------------------------------------------------
def train_model(X, y, model=None):
    """
    Trains a Random Forest model or a provided model.
    
    Arguments:
        X: Feature DataFrame
        y: Target Series
        model: Pre-initialized model (optional)
    
    Returns:
        model: Trained model
        X_test: Test features
        y_test: Test target values
    """
    # Split into train and test sets
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.25, random_state=42, stratify=y
    )
    
    # If no model is provided, use a default RandomForest
    if model is None:
        model = RandomForestClassifier(
            n_estimators=100, 
            class_weight='balanced',
            random_state=42
        )
    
    model.fit(X_train, y_train)
    
    train_acc = model.score(X_train, y_train)
    test_acc = model.score(X_test, y_test)
    
    print(f"✅ Model Training Complete!")
    print(f"Training Accuracy: {train_acc:.4f}")
    print(f"Test Accuracy:     {test_acc:.4f}")
    
    return model, X_test, y_test, X_train, y_train

# -----------------------------------------------------------------------------
# 8) Model Evaluation
# -----------------------------------------------------------------------------
def evaluate_model(model, X_test, y_test, label_encoder, feature_names=None):
    """
    Evaluates the model with multiple metrics and visualizations.
    
    Arguments:
        model: Trained model
        X_test: Test features
        y_test: Test target values
        label_encoder: LabelEncoder for converting numeric predictions to labels
        feature_names: Optional list of feature names for feature importance plot
    
    Returns:
        Dictionary of evaluation metrics
    """
    y_pred = model.predict(X_test)
    y_pred_proba = model.predict_proba(X_test)[:, 1]  # Probability of the positive class
    
    # Convert numeric predictions to labels
    y_test_labels = label_encoder.inverse_transform(y_test)
    y_pred_labels = label_encoder.inverse_transform(y_pred)
    
    # Print metrics
    print("\nClassification Report:")
    print(classification_report(y_test, y_pred))
    
    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred)
    recall = recall_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)
    
    print(f"Accuracy:  {accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall:    {recall:.4f}")
    print(f"F1 Score:  {f1:.4f}")

    # Confusion Matrix
    cm = confusion_matrix(y_test, y_pred)
    plt.figure(figsize=(8, 6))
    disp = ConfusionMatrixDisplay(
        confusion_matrix=cm, 
        display_labels=label_encoder.classes_
    )
    disp.plot(cmap='Blues')
    plt.title("Confusion Matrix")
    plt.tight_layout()
    plt.savefig("confusion_matrix.png")
    print("\nConfusion matrix saved as 'confusion_matrix.png'")
    
    # ROC Curve
    fpr, tpr, _ = roc_curve(y_test, y_pred_proba)
    roc_auc = auc(fpr, tpr)
    
    plt.figure(figsize=(8, 6))
    plt.plot(fpr, tpr, color='darkorange', lw=2, label=f'ROC curve (area = {roc_auc:.2f})')
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Receiver Operating Characteristic (ROC) Curve')
    plt.legend(loc="lower right")
    plt.tight_layout()
    plt.savefig("roc_curve.png")
    print("ROC curve saved as 'roc_curve.png'")
    
    # Feature Importance (top 20 most important features)
    feature_importance = model.feature_importances_
    
    if feature_names is None:
        feature_names = X_test.columns
        
    # Create DataFrame with feature names and importance scores
    feature_importance_df = pd.DataFrame({
        'feature': feature_names,
        'importance': feature_importance
    }).sort_values('importance', ascending=False)
    
    # Save full feature importance to CSV
    feature_importance_df.to_csv('feature_importance.csv', index=False)
    print("Full feature importance saved to 'feature_importance.csv'")
    
    # Plot top 20 features
    plt.figure(figsize=(10, 8))
    sns.barplot(x='importance', y='feature', data=feature_importance_df.head(20))
    plt.title('Top 20 Feature Importance')
    plt.tight_layout()
    plt.savefig("feature_importance.png")
    print("Feature importance plot saved as 'feature_importance.png'")
    
    # Separate feature importance by data source (historical vs current)
    hist_features = [f for f in feature_names if not f.endswith('_current')]
    current_features = [f for f in feature_names if f.endswith('_current')]
    
    # Calculate importance by source
    hist_importance = sum(feature_importance[list(feature_names).index(f)] for f in hist_features if f in feature_names)
    current_importance = sum(feature_importance[list(feature_names).index(f)] for f in current_features if f in feature_names)
    
    total_importance = hist_importance + current_importance
    if total_importance > 0:
        print(f"\nFeature Importance by Data Source:")
        print(f"Historical Features: {hist_importance/total_importance:.1%}")
        print(f"Current Snapshot Features: {current_importance/total_importance:.1%}")
    
    return {
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1_score": f1,
        "confusion_matrix": cm,
        "roc_auc": roc_auc,
        "feature_importance": feature_importance_df
    }

# -----------------------------------------------------------------------------
# 9) Feature Analysis
# -----------------------------------------------------------------------------
def analyze_features(X, y, feature_names, label_encoder):
    """
    Analyzes feature distributions and correlations
    
    Arguments:
        X: Feature DataFrame
        y: Target Series (numeric encoded values)
        feature_names: List of feature column names
        label_encoder: LabelEncoder for trend labels
    """
    # Create a copy of just the features
    data_numeric = X.copy()
    
    # Add the numeric target for correlation
    data_numeric['target'] = y.values
    
    # Print debug info
    print("\nDebug - Data Types:")
    print(data_numeric.dtypes.head())
    
    # Ensure all columns are numeric (filter out any non-numeric)
    numeric_cols = data_numeric.select_dtypes(include=['number']).columns.tolist()
    print(f"Using {len(numeric_cols)} numeric columns for correlation analysis")
    
    # Group features by type (historical vs. current)
    historical_features = [f for f in numeric_cols if not f.endswith('_current') and f != 'target']
    current_features = [f for f in numeric_cols if f.endswith('_current')]
    
    print(f"Feature breakdown:")
    print(f"- Historical/technical features: {len(historical_features)}")
    print(f"- Current snapshot features: {len(current_features)}")
    
    # Correlation with target
    # Extract correlations with target for all features
    correlations = []
    for col in numeric_cols:
        if col != 'target':
            corr = data_numeric[col].corr(data_numeric['target'])
            if not pd.isna(corr):
                correlations.append((col, corr))
    
    # Sort by absolute correlation and get top 20
    top_correlations = sorted(correlations, key=lambda x: abs(x[1]), reverse=True)[:20]
    
    # Create a correlation DataFrame for the top features
    top_corr_features = [item[0] for item in top_correlations]
    top_corr_features.append('target')
    
    # Create correlation matrix for top correlated features
    top_corr_matrix = data_numeric[top_corr_features].corr()
    
    # Correlation Heatmap (top correlated features only)
    plt.figure(figsize=(12, 10))
    try:
        sns.heatmap(top_corr_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)
        plt.title('Correlation Matrix of Top Features')
        plt.tight_layout()
        plt.savefig("top_correlation_matrix.png")
        print("Top correlation matrix saved as 'top_correlation_matrix.png'")
        
        # Save full correlation matrix to CSV
        full_correlation_matrix = data_numeric[numeric_cols].corr()
        full_correlation_matrix.to_csv('full_correlation_matrix.csv')
        print("Full correlation matrix saved to 'full_correlation_matrix.csv'")
    except Exception as e:
        print(f"Warning: Could not create correlation matrix: {str(e)}")
    
    # Create a DataFrame with features and string labels for visualization
    data_viz = X.copy()
    data_viz['next_quarter_trend'] = label_encoder.inverse_transform(y)
    
    # Feature distributions - Select top 6 features based on correlation with target
    top_features = [item[0] for item in top_correlations[:6]]
    
    # If we have top features, plot their distributions
    if top_features:
        plt.figure(figsize=(15, 10))
        for i, feature in enumerate(top_features):
            if i >= 6:  # Limit to 6 features
                break
                
            plt.subplot(2, 3, i+1)
            sns.boxplot(x='next_quarter_trend', y=feature, data=data_viz)
            plt.title(f'Distribution of {feature}')
            
        plt.tight_layout()
        plt.savefig("feature_distributions.png")
        print("Feature distributions saved as 'feature_distributions.png'")
    else:
        print("Warning: Not enough features for meaningful distribution plots")
    
    # Simple feature analysis - print features with highest absolute correlation with target
    print("\nTop 10 Features most correlated with price trend (absolute correlation):")
    for feat, corr in top_correlations[:10]:
        source_type = "Current Snapshot" if feat.endswith("_current") else "Historical/Technical"
        print(f"{feat}: {corr:.4f} [{source_type}]")

# -----------------------------------------------------------------------------
# 10) Make Predictions
# -----------------------------------------------------------------------------
def make_predictions(model, X_predict, label_encoder, tickers, quarter_dates, output_file='stock_trend_predictions.xlsx'):
    """
    Makes predictions for the next quarter's trend
    
    Arguments:
        model: Trained model
        X_predict: Feature DataFrame for prediction
        label_encoder: LabelEncoder to convert numeric predictions to labels
        tickers: List of ticker symbols
        quarter_dates: List of quarter dates
        output_file: Path to save predictions Excel file
        
    Returns:
        DataFrame with predictions
    """
    if X_predict.empty:
        print("Error: X_predict is empty, no predictions can be made")
        return pd.DataFrame()
        
    # Make predictions (both class and probability)
    y_pred = model.predict(X_predict)
    y_pred_proba = model.predict_proba(X_predict)
    
    # Ensure we get the probability for the positive class (up)
    up_idx = list(label_encoder.classes_).index('up') if 'up' in label_encoder.classes_ else 1
    y_pred_proba_up = y_pred_proba[:, up_idx]
    
    # Convert numeric predictions to labels
    y_pred_labels = label_encoder.inverse_transform(y_pred)
    
    # Create results DataFrame
    results = pd.DataFrame({
        'ticker': tickers,
        'quarter_date': quarter_dates,
        'predicted_trend': y_pred_labels,
        'up_probability': y_pred_proba_up.round(4),
        'down_probability': (1 - y_pred_proba_up).round(4)
    })
    
    # Save to Excel with color formatting
    try:
        from openpyxl.styles import PatternFill
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            results.to_excel(writer, index=False)
            
            # Access the worksheet
            worksheet = writer.sheets['Sheet1']
            
            # Define fills
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # Light red
            
            # Find the trend column
            trend_col = 3  # Usually the 3rd column (0-indexed would be 2, but Excel is 1-indexed)
            
            # Apply conditional formatting - start from row 2 to skip header
            for row in range(2, len(results) + 2):
                cell = worksheet.cell(row=row, column=trend_col)
                if cell.value == 'up':
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Predictions saved to '{output_file}' with color highlighting")
        
    except Exception as e:
        # Fallback to basic Excel without formatting
        results.to_excel(output_file, index=False)
        print(f"Predictions saved to '{output_file}' (without color highlighting: {str(e)})")
    
    return results

# -----------------------------------------------------------------------------
# 11) Main Execution
# -----------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Enhanced Stock Price Trend Prediction Model')
    parser.add_argument('--tickers', nargs='+', default=DEFAULT_TICKERS,
                        help=f'Ticker symbols to analyze (default: {DEFAULT_TICKERS})')
    parser.add_argument('--start-date', default=DEFAULT_START_DATE,
                        help=f'Start date for historical data (YYYY-MM-DD) (default: {DEFAULT_START_DATE})')
    parser.add_argument('--end-date', default=DEFAULT_END_DATE,
                        help=f'End date for historical data (YYYY-MM-DD) (default: {DEFAULT_END_DATE})')
    parser.add_argument('--model-file', default=DEFAULT_MODEL_FILE,
                        help=f'Path to save/load model (default: {DEFAULT_MODEL_FILE})')
    parser.add_argument('--predict-only', action='store_true',
                        help='Only make predictions using an existing model')
    parser.add_argument('--feature-importance', action='store_true',
                        help='Only analyze feature importance of an existing model')
    parser.add_argument('--historical-only', action='store_true',
                        help='Only use historical and technical data (ignore current snapshot)')
    
    args = parser.parse_args()
    
    print(f"=== Enhanced Stock Price Trend Prediction Model ===")
    print(f"Tickers: {args.tickers}")
    print(f"Date Range: {args.start_date} to {args.end_date}")
    
    # Feature importance analysis mode
    if args.feature_importance and os.path.exists(args.model_file):
        # Load model
        model, scaler, feature_names, label_encoder, success = load_saved_model(args.model_file)
        if not success:
            print("Error loading model. Exiting feature importance mode.")
            return
            
        # Create a minimal dataset for analysis
        print("\nAnalyzing feature importance of the model...")
        
        # Get feature importance directly from the model
        importances = model.feature_importances_
        
        # Create and sort feature importance DataFrame
        importance_df = pd.DataFrame({
            'Feature': feature_names,
            'Importance': importances
        }).sort_values('Importance', ascending=False)
        
        # Save full feature importance to CSV
        importance_df.to_csv('feature_importance.csv', index=False)
        print("Feature importance saved to 'feature_importance.csv'")
        
        # Plot top 20 features
        plt.figure(figsize=(12, 10))
        sns.barplot(x='Importance', y='Feature', data=importance_df.head(20))
        plt.title('Top 20 Most Important Features')
        plt.tight_layout()
        plt.savefig("feature_importance.png")
        print("Feature importance plot saved as 'feature_importance.png'")
        
        # Print top 10 features
        print("\nTop 10 most important features:")
        for i, (feature, importance) in enumerate(zip(importance_df['Feature'][:10], importance_df['Importance'][:10])):
            feature_type = "Current" if feature.endswith("_current") else "Historical/Technical"
            print(f"{i+1}. {feature}: {importance:.4f} [{feature_type}]")
        
        # Group features by type
        hist_features = [f for f in feature_names if not f.endswith('_current')]
        current_features = [f for f in feature_names if f.endswith('_current')]
        
        # Calculate importance by source
        hist_importance = sum(importances[list(feature_names).index(f)] for f in hist_features if f in feature_names)
        current_importance = sum(importances[list(feature_names).index(f)] for f in current_features if f in feature_names)
        
        total_importance = hist_importance + current_importance
        if total_importance > 0:
            print(f"\nFeature Importance by Data Source:")
            print(f"Historical/Technical Features: {hist_importance/total_importance:.1%}")
            print(f"Current Snapshot Features: {current_importance/total_importance:.1%}")
            
        return
    
    # If predict-only mode and model exists, load it
    if args.predict_only and os.path.exists(args.model_file):
        # Load model
        model, scaler, feature_names, label_encoder, success = load_saved_model(args.model_file)
        if not success:
            print("Error loading model. Exiting prediction-only mode.")
            return
        
        # Fetch data for prediction
        stock_data = fetch_stock_data(args.tickers, args.start_date, args.end_date)
        quarterly_prices = get_quarterly_prices(stock_data)
        financial_data = fetch_financial_ratios(args.tickers)
        
        # Extract the most recent quarter for each ticker using the new extraction function
        combined_data = extract_quarterly_ratios(financial_data, quarterly_prices)
        
        if combined_data.empty:
            print("No valid prediction data could be prepared. Exiting.")
            return
            
        # Get only the last quarter for each ticker
        prediction_data = []
        for ticker in args.tickers:
            ticker_data = combined_data[combined_data['ticker'] == ticker]
            if not ticker_data.empty:
                # Get the latest quarter
                latest_quarter = ticker_data.loc[ticker_data['quarter_date'].idxmax()]
                prediction_data.append(latest_quarter)
        
        # Convert to DataFrame
        prediction_df = pd.DataFrame(prediction_data)
        
        # If historical-only mode, filter out current snapshot features
        if args.historical_only:
            print("Using historical and technical features only (ignoring current snapshot)")
            prediction_df = prediction_df[[col for col in prediction_df.columns if not col.endswith('_current')]]
        
        # Prepare prediction data
        # Create a minimal training dataset to ensure proper processing
        dummy_train = pd.DataFrame({
            'ticker': ['DUMMY'],
            'quarter_date': [pd.Timestamp.now()],
            'price': [100.0],
            'quarterly_return': [0.0],
            'trend': ['up'],
            'next_quarter_trend': ['up'],
            'data_source': ['historical']
        })
        
        # Add dummy values for all required features
        for col in feature_names:
            if col not in dummy_train.columns:
                dummy_train[col] = 0.0
                
        # Add next_quarter_trend to prediction_df for the preprocessing function if needed
        if 'next_quarter_trend' not in prediction_df.columns:
            dummy_y = pd.Series(['up'] * len(prediction_df))
            prediction_df['next_quarter_trend'] = dummy_y
        
        # Use preprocessing function with prediction_df as the second argument
        _, _, X_predict, _, _, _ = preprocess_data(dummy_train, prediction_df)
        
        # Check if X_predict is valid
        if X_predict.empty:
            print("Warning: X_predict is empty, creating empty DataFrame with required columns")
            X_predict = pd.DataFrame(columns=feature_names)
        
        # Ensure all required features are present and in the right order
        for col in feature_names:
            if col not in X_predict.columns:
                print(f"Adding missing feature: {col}")
                X_predict[col] = 0
                
        # Reorder columns to match training data
        try:
            X_predict = X_predict[feature_names]
        except Exception as e:
            print(f"Error reordering columns: {str(e)}")
            # Create a new DataFrame with the right structure
            X_predict_fixed = pd.DataFrame(columns=feature_names)
            for col in feature_names:
                if col in X_predict.columns:
                    X_predict_fixed[col] = X_predict[col]
                else:
                    X_predict_fixed[col] = 0
            X_predict = X_predict_fixed
        
        # Make predictions
        predictions = make_predictions(
            model, 
            X_predict, 
            label_encoder, 
            prediction_df['ticker'].values,
            prediction_df['quarter_date'].values,
            output_file='next_quarter_predictions.xlsx'
        )
        
        if not predictions.empty:
            print("\nSample predictions:")
            print(predictions)
        else:
            print("No predictions could be generated.")
        
    else:
        # Full training workflow
        # 1. Fetch stock data
        stock_data = fetch_stock_data(args.tickers, args.start_date, args.end_date)
        quarterly_prices = get_quarterly_prices(stock_data)
        
        # 2. Fetch financial ratios
        financial_data = fetch_financial_ratios(args.tickers)
        
        # 3. Extract and combine quarterly data with the new function
        combined_data = extract_quarterly_ratios(financial_data, quarterly_prices)
        
        if combined_data.empty:
            print("No valid combined data could be created. Exiting.")
            return
            
        # If historical-only mode, filter out current snapshot features
        if args.historical_only:
            print("Using historical and technical features only (ignoring current snapshot)")
            combined_data = combined_data[[col for col in combined_data.columns if not col.endswith('_current')]]
        
        # Save raw data to Excel for inspection
        combined_data.to_excel('raw_quarterly_data.xlsx', index=False)
        print("Raw quarterly data saved to 'raw_quarterly_data.xlsx'")
        
        # 4. Preprocess data
        X_train, y_train, _, feature_names, label_encoder, scaler = preprocess_data(combined_data)
        
        print(f"Training features shape: {X_train.shape}")
        print(f"Training target shape: {y_train.shape}")
        
        # 5. Feature Analysis
        print("\n=== Feature Analysis ===")
        analyze_features(X_train, y_train, feature_names, label_encoder)
        
        # 6. Cross-Validation
        print("\n=== Cross-Validation (Baseline RandomForest) ===")
        cross_validate_model(X_train, y_train, n_splits=5)
        
        # 7. Hyperparameter Tuning
        print("\n=== Hyperparameter Tuning with Grid Search ===")
        best_model_grid = grid_search_tuning(X_train, y_train)
        
        print("\n=== Hyperparameter Tuning with Optuna ===")
        best_model_optuna = tune_rf_with_optuna(X_train, y_train, n_trials=20)
        
        # Compare models
        print("\n=== Comparing Grid Search and Optuna Models ===")
        skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        
        grid_scores = cross_val_score(best_model_grid, X_train, y_train, cv=skf, scoring='f1')
        optuna_scores = cross_val_score(best_model_optuna, X_train, y_train, cv=skf, scoring='f1')
        
        print(f"Grid Search model mean F1:   {grid_scores.mean():.4f}")
        print(f"Optuna model mean F1:        {optuna_scores.mean():.4f}")
        
        # Choose the better model
        if optuna_scores.mean() > grid_scores.mean():
            print("\nOptuna model is better. Using Optuna model for final training.")
            chosen_model = best_model_optuna
        else:
            print("\nGrid Search model is better. Using Grid Search model for final training.")
            chosen_model = best_model_grid
        
        # 8. Train Final Model
        print("\n=== Training Final Model ===")
        final_model, X_test, y_test, _, _ = train_model(X_train, y_train, model=chosen_model)
        
        # 9. Evaluate Model
        print("\n=== Final Model Evaluation ===")
        eval_results = evaluate_model(final_model, X_test, y_test, label_encoder, feature_names)
        
        # 10. Save the model
        with open(args.model_file, 'wb') as f:
            pickle.dump({
                'model': final_model,
                'scaler': scaler,
                'feature_names': feature_names,
                'label_encoder': label_encoder  # Save the label encoder too
            }, f)
        print(f"\nModel saved to '{args.model_file}'")
        
        # 11. Make predictions for the next quarter
        # Get the most recent quarter for each ticker
        prediction_data = []
        for ticker in args.tickers:
            ticker_data = combined_data[combined_data['ticker'] == ticker]
            if not ticker_data.empty:
                # Get the latest quarter
                latest_quarter = ticker_data.loc[ticker_data['quarter_date'].idxmax()]
                prediction_data.append(latest_quarter)
        
        # Convert to DataFrame
        prediction_df = pd.DataFrame(prediction_data)
            
        # Use the combined_data as training reference and prediction_df as prediction data
        _, _, X_predict, _, _, _ = preprocess_data(combined_data, prediction_df)
        
        # Check if X_predict is valid
        if X_predict.empty:
            print("Warning: X_predict is empty, creating DataFrame with required columns")
            X_predict = pd.DataFrame(columns=feature_names)
        
        # Ensure all features are present
        for col in feature_names:
            if col not in X_predict.columns:
                print(f"Adding missing feature: {col}")
                X_predict[col] = 0
        
        # Ensure feature order matches training
        try:
            X_predict = X_predict[feature_names]
            print(f"Final prediction data shape: {X_predict.shape}")
        except Exception as e:
            print(f"Error reordering prediction columns: {str(e)}")
            # Create a new DataFrame with the correct structure
            X_predict_fixed = pd.DataFrame(columns=feature_names)
            for col in feature_names:
                if col in X_predict.columns:
                    X_predict_fixed[col] = X_predict[col]
                else:
                    X_predict_fixed[col] = 0
            X_predict = X_predict_fixed
        
        # Make predictions for the next quarter
        next_quarter = (pd.to_datetime(prediction_df['quarter_date'].iloc[0]) + pd.DateOffset(months=3)).strftime('%Y-%m-%d')
        print(f"\n=== Making Predictions for Next Quarter (around {next_quarter}) ===")
        
        predictions = make_predictions(
            final_model, 
            X_predict, 
            label_encoder, 
            prediction_df['ticker'].values,
            prediction_df['quarter_date'].values
        )
        
        if not predictions.empty:
            print("\nNext quarter predictions:")
            print(predictions)
        else:
            print("No predictions could be generated.")

if __name__ == "__main__":
    main()